import mysql.connector
import argparse
import openpyxl
import sys
import re
import os


def db_check(args, conn) -> str:

    SQL = 'USE '

    if args.d is not None:
        SQL += str(args.d)

        cur = conn.cursor()

        try:
            cur.execute(SQL)
        except Exception as e:
            CREATE_DB = 'CREATE DATABASE {}'.format(args.d)
            cur.execute(CREATE_DB)
            print('[-] ' + str(sys.exc_info()[0] ) + str(e))
        
        conn.commit()

        return args.d

    else:
        pattern = re.compile('\w*.xlsx')
        file_name = pattern.findall(args.f)[0] .split('.')[0] 

        CREATE_DB = 'CREATE DATABASE {}'.format(file_name)
        
        cur = conn.cursor()

        try:
            cur.execute(CREATE_DB)
            print('[+] ' + str(CREATE_DB))
        except Exception as e:
            print('[-] ' + str(sys.exc_info()[0] ) + str(e))

        conn.commit()

        return file_name


def row_check(sheet, head_list) -> int:

    rows = 2
    columns = 1

    while True:
        check = 0

        for idx, i in enumerate(range(len(head_list))):
            if sheet.cell(row=rows, column=columns).value == None:
                check += 1
            columns += 1

        if check == len(head_list):
            break
        else:
            rows += 1
            columns = 1
    
    return rows - 2


def excel_parse(args, conn, DB_name) -> int:
   
    wb = openpyxl.load_workbook(args.f)
    sheet = wb.active
    sheet_name = wb.sheetnames

    head_list = list()
    data_type = list()
    
    rows = 1
    columns = 1
    
    if (sheet.cell(row=1, column=1).value == None):
        return 1
    else:
        while sheet.cell(row=1, column=columns).value != None:
            head_list.append(sheet.cell(row=1, column=columns).value)
            columns += 1           
    print(head_list)

    row_size = row_check(sheet, head_list)
    print(row_size)

    rows = 2 
    columns = 1
    for idx, i in enumerate(range(len(head_list))):
        int_check = 1 
        for j in range(row_size):
            try:
                if sheet.cell(row=rows, column=columns).value == None:
                    int_check = 1
                else:
                    int(sheet.cell(row=rows, column=columns).value)
            except Exception as e:
                int_check = 0
                break
            rows += 1

        if int_check == 1:
            data_type.append('int')
        else:
            max_size = 256
            for j in range(row_size):
                if max_size < len(sheet.cell(row=rows, column=columns).value) \
                    and sheet.cell(row=rows, column=columns).value != None:
                    max_size = len(sheet.cell(row=rows, column=columns).value)

            type_string = 'varchar({})'.format(int(max_size*1.2))
            data_type.append(type_string) 

        rows = 2
        columns += 1
    print(data_type)

    if args.t is not None:
        create_table(conn, DB_name, args.t, head_list, data_type)
        insert_value(conn, sheet, DB_name, args.t, head_list, row_size, data_type)
    else:
        create_table(conn, DB_name, sheet_name[0], head_list, data_type)
        insert_value(conn, sheet, DB_name, sheet_name[0], head_list, row_size, data_type)

    wb.close()
    
    return 0


def create_table(conn, DB_name, table_name, head_list, data_type) -> None:
    CREATE_TABLE_SQL = 'CREATE TABLE {}.{} ('.format(DB_name, table_name)

    cur = conn.cursor()
    
    for idx, head in enumerate(head_list):
        if idx == (len(head_list)-1):
            CREATE_TABLE_SQL += '{} {})'.format(head, data_type[idx])
        else:
            CREATE_TABLE_SQL += '{} {}, '.format(head, data_type[idx])

    try:
        cur.execute(CREATE_TABLE_SQL)
        print('[+] ' + str(CREATE_TABLE_SQL))
    except Exception as e:
        print('[-] ' + str(sys.exc_info()[0]) + str(e))

    conn.commit()


def insert_value(conn, sheet, DB_name, table_name, head_list, row_size, data_type) -> None:
    
    rows = 2
    columns = 1

    cur = conn.cursor()

    for type_idx, i in enumerate(range(row_size)):
        INSERT_SQL = 'INSERT INTO {}.{} VALUES('.format(DB_name, table_name)
        for idx, j in enumerate(range(len(head_list))):
            if data_type[idx] == 'int':
                if sheet.cell(row=rows, column=columns).value == None:
                    if idx == (len(head_list) - 1):
                        INSERT_SQL += '0)'
                    else:
                        INSERT_SQL += '0, '
                else:
                    if idx == (len(head_list) - 1):
                        INSERT_SQL += '{})'.format(sheet.cell(row=rows, column=columns).value)
                    else:
                        INSERT_SQL += '{}, '.format(sheet.cell(row=rows, column=columns).value)

            else:
                if sheet.cell(row=rows, column=columns).value == None:
                    if idx == (len(head_list) - 1):
                        INSERT_SQL += '\'\')'
                    else:
                        INSERT_SQL += '\'\', '
                else:
                    if idx == (len(head_list) - 1):
                        INSERT_SQL += '\'{}\')'.format(sheet.cell(row=rows, column=columns).value)
                    else:
                        INSERT_SQL += '\'{}\', '.format(sheet.cell(row=rows, column=columns).value)
            columns += 1

        try:
            cur.execute(INSERT_SQL)
            print('[+] ' + str(INSERT_SQL))
        except Exception as e:
            print('[-] ' + str(sys.exc_info()[0]) + str(e))
        rows += 1
        columns = 1

    conn.commit()
    

def mysql_connect(args):

    try:
        conn = mysql.connector.connect(
            host = args.host,
            port = args.port,
            user = args.u,
            passwd = args.p
        ) 
    except Exception as e:
        print('Connect Error')
   
    DB_name = db_check(args, conn) 
    excel_parse(args, conn, DB_name)
    
    conn.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--host', help='host', required=True)
    parser.add_argument('--port', help='Default 3306', default=3306, required=False)
    parser.add_argument('-u', help='User', required=True)
    parser.add_argument('-p', help='Password', required=True)
    parser.add_argument('-d', help='DB Name(Default file name)', required=False)
    parser.add_argument('-t', help='Table Name(Default Sheet name)', required=False)
    parser.add_argument('-f', help='excel file path', required=True)

    args = parser.parse_args()

    mysql_connect(args)
